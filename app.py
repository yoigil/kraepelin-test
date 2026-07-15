import streamlit as st
import random
import time
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
from zipencrypt import ZipFile

try:
    from streamlit_autorefresh import st_autorefresh
    HAS_AUTOREFRESH = True
except ImportError:
    HAS_AUTOREFRESH = False

# --- App Settings ---
TOTAL_COLUMNS = 50
TIME_PER_COLUMN = 15

st.set_page_config(page_title="Tes Kraepelin", layout="centered")

# --- Initialize Global Session Variables Robustly ---
defaults = {
    "step": "login",
    "user_name": "",
    "user_nik": "",
    "current_column": 0,
    "row_index": 0,
    "start_time": 0.0,
    "attempts": [0] * TOTAL_COLUMNS,
    "corrects": [0] * TOTAL_COLUMNS,
    "errors": [0] * TOTAL_COLUMNS,
    "numbers_matrix": [[random.randint(1, 9) for _ in range(120)] for _ in range(TOTAL_COLUMNS)]
}

for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val

# --- Helper Functions ---
def next_column():
    st.session_state["current_column"] += 1
    st.session_state["row_index"] = 0
    st.session_state["start_time"] = time.time()
    if st.session_state["current_column"] >= TOTAL_COLUMNS:
        st.session_state["step"] = "finished"

def process_answer(user_input, correct_digit):
    col = st.session_state["current_column"]
    st.session_state["attempts"][col] += 1

    if user_input == correct_digit:
        st.session_state["corrects"][col] += 1
    else:
        st.session_state["errors"][col] += 1

    st.session_state["row_index"] += 1

def generate_zipped_xlsx():
    wb = openpyxl.Workbook()

    # Sheet 1: Raw Data
    ws_data = wb.active
    ws_data.title = "Data Per Kolom"
    ws_data.views.sheetView[0].showGridLines = True
    ws_data.append(["Kolom", "Jumlah Benar", "Jumlah Salah", "Total Jumlah"])

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for c_num in range(1, 5):
        cell = ws_data.cell(row=1, column=c_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for i in range(TOTAL_COLUMNS):
        ws_data.append([i + 1, st.session_state["corrects"][i], st.session_state["errors"][i], st.session_state["attempts"][i]])

    # Sheet 2: Summary Dashboard (Ringkasan Hasil)
    ws_sum = wb.create_sheet(title="Ringkasan Hasil", index=0)
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum["A1"] = "Laporan Hasil Tes Kraepelin"
    ws_sum["A1"].font = Font(name="Arial", size=16, bold=True, color="1F497D")
    ws_sum["A2"] = f"Nama Peserta: {st.session_state.get('user_name', '')}"
    ws_sum["A2"].font = Font(name="Arial", size=11, bold=True)
    ws_sum["A3"] = f"NIK Peserta: {st.session_state.get('user_nik', '')}"
    ws_sum["A3"].font = Font(name="Arial", size=11, bold=True)

    headers = ["Points", "Score", "Criteria"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=5, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    metrics = [
        (
            "PANKER (Kecepatan)",
            "=SUM('Data Per Kolom'!D2:D51)/MAX('Data Per Kolom'!A2:A51)",
            '=IF(B6<11, "Rendah", IF(B6>14.9, "Tinggi", "Sedang"))'
        ),
        (
            "TIANKER (Ketelitian)",
            "=SUM('Data Per Kolom'!C2:C51)",
            '=IF(B7>=23, "Rendah", IF(B7<=5, "Tinggi", "Sedang"))'
        ),
        (
            "JANKER (Keajegan)",
            "=MAX('Data Per Kolom'!D2:D51)-MIN('Data Per Kolom'!D2:D51)",
            '=IF(B8<9, "Tinggi", IF(B8>12, "Rendah", "Sedang"))'
        ),
        (
            "HANKER (Ketahanan)",
            "=50*SLOPE('Data Per Kolom'!D2:D51,'Data Per Kolom'!A2:A51)",
            '=IF(B9<-1.947, "Rendah", IF(B9>0.223, "Tinggi", "Sedang"))'
        )
    ]

    for idx, (label, score_formula, criteria_formula) in enumerate(metrics, start=6):
        ws_sum.cell(row=idx, column=1, value=label).font = Font(name="Arial", bold=True)
        ws_sum.cell(row=idx, column=2, value=score_formula).alignment = center_align
        ws_sum.cell(row=idx, column=3, value=criteria_formula).alignment = center_align

    chart = LineChart()
    chart.title = "Grafik Ketahanan Kerja (Hanker)"
    chart.width = 15
    chart.height = 8
    chart.legend = None
    chart.add_data(Reference(ws_data, min_col=4, min_row=1, max_row=51), titles_from_data=True)
    chart.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=51))

    ws_sum.add_chart(chart, "A12")

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    zip_buffer = io.BytesIO()

    clean_name = st.session_state.get('user_name', 'peserta').replace(' ', '_')
    internal_excel_name = f"Laporan_Kraepelin_{clean_name}.xlsx"
    password = b"HR3918"

    with ZipFile(zip_buffer, mode="w") as zf:
        zf.writestr(internal_excel_name, excel_buffer.getvalue(), pwd=password)

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# --- Single placeholder that owns ALL rendered content ---
# Rendering every view inside the SAME st.empty() container (and fully
# overwriting it each rerun) is what prevents leftover widgets from a
# previous view lingering on screen when the step changes.
main_area = st.empty()

current_step = st.session_state.get("step", "login")

# --- VIEW 1: Registration ---
if current_step == "login":
    with main_area.container():
        st.title("Psikotest Kraepelin")
        st.divider()

        st.warning(
            "**PERINGATAN PENTING:** Tes ini menggunakan penghitung waktu otomatis. "
            "Setiap kolom berjalan selama tepat **15 detik**. Ketika waktu habis, sistem akan "
            "memindahkan Anda ke kolom berikutnya secara paksa. **Jangan refresh halaman selama tes berlangsung!**",
            icon="⚠️"
        )

        st.warning(
            "**NOTE:** Apabila mengerjakan tes ini melalui smartphone, "
            "**Tolong gunakan Google Chrome dan pastikan website sudah dalam Mode Website**.  "
            "Tekan titik tiga pada pojok kanan atas, dan cek mode website.",
            icon="⚠️"
        )
        st.divider()

        st.markdown("### Panduan Pelaksanaan Tes:")

        col_g1, col_g2 = st.columns(2)
        with col_g1:
            st.markdown(
                "**Cara Menjawab:**\n"
                "1. Anda akan melihat **dua angka aktif** berwarna biru di tengah layar.\n"
                "2. **Jumlahkan kedua angka** tersebut (Angka Atas + Angka Bawah).\n"
                "3. Ambil **angka digit terakhir** saja dari hasil penjumlahan.\n"
                "   * *Contoh:* $4 + 5 = \\mathbf{9}$ (Ketik **9**)\n"
                "   * *Contoh:* $8 + 7 = 1\\mathbf{5}$ (Ketik **5**)"
            )
        with col_g2:
            st.markdown(
                "**Metode Input & Kontrol:**\n"
                "* **Layar Sentuh/Mouse:** Anda bisa mengklik tombol angka yang tertera pada numpad digital di layar.\n"
                "* Tes terdiri dari **50 Kolom berturut-turut**."
            )

        st.divider()
        st.markdown("### Identitas Peserta")

        # A form batches the two text inputs + button into a single atomic
        # submission, so no rerun happens until "Mulai Tes!" is actually
        # clicked (typing no longer triggers intermediate reruns).
        with st.form("login_form"):
            name_input = st.text_input(
                "Masukkan Nama Lengkap Anda sesuai KTP dengan huruf kapital:",
                value=st.session_state.get("user_name", ""),
                placeholder="Contoh: BUDI SANTOSO"
            )
            nik_input = st.text_input(
                "Masukkan NIK KTP anda:",
                value=st.session_state.get("user_nik", ""),
                placeholder="Contoh: 3171xxxxxxxxxxxx"
            )
            submitted = st.form_submit_button("Mulai Tes!", type="primary")

        if submitted:
            if not name_input.strip() or not nik_input.strip():
                st.error("Nama dan NIK tidak boleh kosong!")
            else:
                st.session_state["user_name"] = name_input.strip()
                st.session_state["user_nik"] = nik_input.strip()
                st.session_state["step"] = "active_test"
                st.session_state["start_time"] = time.time()
                st.rerun()

# --- VIEW 2: Active Kraepelin Evaluation ---
elif current_step == "active_test":
    with main_area.container():
        # Re-run this whole script once per second (instead of 10x/sec)
        # purely to update the countdown. This is far less likely to race
        # against a user's button click than the old sleep(0.1)+rerun loop.
        if HAS_AUTOREFRESH:
            st_autorefresh(interval=1000, key="timer_refresh")

        with st.expander("PANDUAN SINGKAT CARA MENJAWAB (Klik untuk tutup/buka)", expanded=True):
            st.info(
                "**Jumlahkan dua angka aktif** (berwarna biru). Input **digit terakhir** dari hasil penjumlahan.\n\n"
                "* *Contoh:* $4 + 5 = \\mathbf{9}$ (Ketik **9**) | $8 + 7 = 1\\mathbf{5}$ (Ketik **5**)\n"
                "* Gunakan **Keyboard/Numpad** fisik Anda atau klik tombol di bawah untuk menjawab."
            )

        col_idx = st.session_state.get("current_column", 0)
        row_idx = st.session_state.get("row_index", 0)

        elapsed_time = time.time() - st.session_state.get("start_time", time.time())
        time_left = max(0, int(TIME_PER_COLUMN - elapsed_time))

        if elapsed_time >= TIME_PER_COLUMN:
            next_column()
            st.rerun()

        st.subheader(f"Peserta: {st.session_state.get('user_name', '')}")
        c1, c2, c3 = st.columns(3)
        c1.metric("Kolom Aktif", f"{col_idx + 1} / {TOTAL_COLUMNS}")
        c2.metric("Sisa Waktu Kolom", f"{time_left} detik")
        c3.metric("Total Soal Terjawab", sum(st.session_state.get("attempts", [0])))

        st.divider()

        matrix = st.session_state.get("numbers_matrix")[col_idx]
        # Guard against running past the end of the pre-generated matrix
        # if someone answers extremely fast.
        max_start = len(matrix) - 4
        if row_idx > max_start:
            row_idx = max_start
            st.session_state["row_index"] = row_idx

        num_bottom_preview = matrix[row_idx]
        num_lower_active = matrix[row_idx + 1]
        num_upper_active = matrix[row_idx + 2]
        num_top_preview = matrix[row_idx + 3]
        correct_digit = (num_lower_active + num_upper_active) % 10

        m1, m2 = st.columns([2, 3])

        with m1:
            st.markdown("### Angka")
            st.markdown(f"<div style='background-color:#f0f2f6; padding:10px; border-radius:5px; text-align:center; font-size:28px; color:#555;'>{num_top_preview}</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='background-color:#1f497d; padding:15px; margin:5px 0; border-radius:5px; text-align:center; font-size:36px; font-weight:bold; color:white;'>{num_upper_active}</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='background-color:#1f497d; padding:15px; margin:5px 0; border-radius:5px; text-align:center; font-size:36px; font-weight:bold; color:white;'>{num_lower_active}</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='background-color:#f0f2f6; padding:10px; border-radius:5px; text-align:center; font-size:28px; color:#555;'>{num_bottom_preview}</div>", unsafe_allow_html=True)

        with m2:
            st.markdown("### Jawaban")

            st.html("""
                <style>
                    div[data-testid="stButton"] button p {
                        font-size: 36px !important;
                        font-weight: bold !important;
                    }
                </style>
            """)

            numpad_layout = [["7", "8", "9"], ["4", "5", "6"], ["1", "2", "3"]]
            for row in numpad_layout:
                btn_cols = st.columns(3)
                for i, num_str in enumerate(row):
                    if btn_cols[i].button(num_str, key=f"btn_{num_str}", use_container_width=True):
                        process_answer(int(num_str), correct_digit)
                        st.rerun()

            if st.button("0", key="btn_0", use_container_width=True):
                process_answer(0, correct_digit)
                st.rerun()

        # Streamlit doesn't expose the widget `key` as a DOM attribute, so
        # matching on button text content is what actually works.
        js_listener = """
        <script>
        const handleKeyDown = (e) => {
            if (e.key >= '0' && e.key <= '9') {
                const buttons = window.parent.document.querySelectorAll('div[data-testid="stButton"] button');
                for (const btn of buttons) {
                    if (btn.innerText.trim() === e.key) {
                        btn.click();
                        break;
                    }
                }
            }
        };
        window.parent.document.removeEventListener('keydown', handleKeyDown);
        window.parent.document.addEventListener('keydown', handleKeyDown);
        </script>
        """
        st.components.v1.html(js_listener, height=0, width=0)

        if not HAS_AUTOREFRESH:
            # Fallback only — much slower than the autorefresh path above,
            # and more prone to the exact race condition you hit. Install
            # `streamlit-autorefresh` (add it to requirements.txt) to avoid
            # relying on this branch at all.
            time.sleep(1)
            st.rerun()

# --- VIEW 3: Final Download Trigger ---
elif current_step == "finished":
    with main_area.container():
        st.balloons()
        st.title("Tes Selesai!")
        st.success(f"Selamat {st.session_state.get('user_name', '')}, Anda telah menyelesaikan seluruh rangkaian ujian Kraepelin!")

        excel_data = generate_zipped_xlsx()

        clean_nik = str(st.session_state.get('user_nik', 'peserta')).strip().replace(' ', '_')

        st.download_button(
            label="Download ZIP Hasil Evaluasi",
            data=excel_data,
            file_name=f"Laporan_Kraepelin_{clean_nik}.zip",
            mime="application/zip",
            type="primary"
        )

        if st.button("Ulangi Tes Baru"):
            st.session_state.clear()
            st.rerun()
